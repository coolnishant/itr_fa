"""
Excel export for Schedule FA Section A3.

Generates a formatted .xlsx file matching the ITR A3 table layout.
Uses openpyxl for Excel generation.
"""

import logging
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column definitions
A3_COLUMNS = [
    ("Sl.\nNo.", 6),
    ("Country\nName and\nCode", 22),
    ("Name of\nentity", 22),
    ("Address of\nentity", 30),
    ("Zip code", 10),
    ("Nature of\nentity", 12),
    ("Date of\nacquiring\nthe interest", 14),
    ("Initial value\nof the\ninvestment", 16),
    ("Peak value of\ninvestment\nduring the\nperiod", 16),
    ("Closing\nbalance", 16),
    ("Total gross\namount\npaid/credited\nw.r.t. holding\nduring the\nperiod", 18),
    ("Total gross\nproceeds from\nsale or\nredemption of\ninvestment\nduring the\nperiod", 18),
]


def _format_inr(value):
    """Format a number in Indian comma style (e.g., 6,20,875)."""
    if value is None or value == 0:
        return "0"
    s = str(int(abs(value)))
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]
    return ("-" if value < 0 else "") + result


def export_a3_excel(rows: list, calendar_year: int, output_path: str = None) -> bytes:
    """
    Generate a formatted Excel file with A3 table data.

    Args:
        rows: List of calculated A3 row dicts from calculator.calculate_a3_rows()
        calendar_year: The calendar year being reported
        output_path: Optional file path to save. If None, returns bytes.

    Returns:
        Excel file as bytes (for download)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule FA - A3 (CY{calendar_year})"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    number_alignment = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    col_num_font = Font(name="Calibri", bold=True, size=9, color="666666")
    col_num_fill = PatternFill(start_color="E8ECF1", end_color="E8ECF1", fill_type="solid")

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"A3 - Details of Foreign Equity and Debt Interest held (including any beneficial interest) in any entity at any time during the calendar year ending as on 31st December, {calendar_year}"
    title_cell.font = Font(name="Calibri", bold=True, size=11)
    title_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 45

    # Column headers (row 2)
    for col_idx, (header, width) in enumerate(A3_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 80

    # Column numbers row (row 3)
    for col_idx in range(1, 13):
        cell = ws.cell(row=3, column=col_idx, value=str(col_idx))
        cell.font = col_num_font
        cell.fill = col_num_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(rows, start=4):
        values = [
            row_data.get("sl_no", ""),
            row_data.get("country", ""),
            row_data.get("entity_name", ""),
            row_data.get("address", ""),
            row_data.get("zip", ""),
            row_data.get("nature", ""),
            row_data.get("acquire_date", ""),
            _format_inr(row_data.get("initial_value")),
            _format_inr(row_data.get("peak_value")),
            _format_inr(row_data.get("closing_balance")),
            _format_inr(row_data.get("total_dividends")),
            _format_inr(row_data.get("sale_proceeds")),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            if col_idx >= 8:  # Number columns
                cell.alignment = number_alignment
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = data_alignment

        ws.row_dimensions[row_idx].height = 30

    # Save
    if output_path:
        wb.save(output_path)
        logger.info(f"Excel saved to {output_path}")

    # Always return bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
